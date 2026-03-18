#!/usr/bin/env python3
"""
iSeq Excel to HTML Variant Report Converter

Converts Combine.xlsx from iSeq pipeline to interactive HTML reports.
Matches the processVCF.sh HTML report style with colored panels, ACMG chips,
and gene hero sections.

Column Structure (based on iSeq Combine.xlsx):
- Col 1: SAMPLE
- Col 2-19: Basic variant information
- Col 20-21: Sample Comparison Data
- Col 22-30: Additional Variant Information
- Col 31-88: Population Frequency Database
- Col 89-122: ACMG Criteria & ClinVar
- Col 123-152: Computational Predictions

Usage:
    python3 excel_to_html_report.py <input_excel_file> [output_directory] [snapshots_directory]
"""

import os
import sys
import re
from pathlib import Path
from typing import Dict, List, Any, Optional
from html import escape
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Key column names: logical key -> xlsx header name
KEY_COL_NAMES = {
    "sample": "SAMPLE", "chr": "Chr", "pos": "Pos", "ref": "Ref", "alt": "Alt", "gt": "GT",
    "gene": "GENE", "transcript": "Transcript", "hgvsg": "HGVSg", "hgvsc": "HGVSc", "hgvsp": "HGVSp",
    "ad": "AD", "dp": "DP", "qual": "QUAL", "vaf": "VAF",
}

# Column group header names (defines display order within each panel)
COLUMN_GROUP_NAMES = {
    "basic_info": [
        "Chr", "Pos", "Ref", "Alt", "GT", "GENE", "Transcript", "HGVSg", "HGVSc", "HGVSp",
        "AD", "DP", "QUAL", "VAF", "snp141",
        "GENE:Chr(GRCh38):HGVSg; Transcript(GENE):HGVSc;HGVSp(0% VAF)",
        "cosmic91", "CancerVar and Evidence",
    ],
    "sample_comparison": ["iSeq305", "TMSP1024"],
    "additional_info": [
        "FILTER", "ExonicFunc.refGene", "Func.ensGene", "cytoBand",
        "Consequence", "STRAND", "VARIANT_CLASS", "EXON", "INTRON",
    ],
    "population_freq": [
        "AF_All", "AF_CHS", "AF_INS", "AF_MAS",
        "AF_SEA", "AF_NEA", "AF_SAS",
        "esp6500siv2_all",
        "ExAC_ALL", "ExAC_AFR", "ExAC_AMR", "ExAC_EAS", "ExAC_FIN", "ExAC_NFE", "ExAC_OTH", "ExAC_SAS",
        "1000g2015aug_all", "1000g2015aug_afr", "1000g2015aug_eas",
        "1000g2015aug_amr", "1000g2015aug_eur", "1000g2015aug_sas",
        "Kaviar_AF", "Kaviar_AC", "Kaviar_AN",
        "gnomad41_exome_AF", "gnomad41_exome_AF_afr", "gnomad41_exome_AF_sas",
        "gnomad41_exome_AF_amr", "gnomad41_exome_AF_eas", "gnomad41_exome_AF_nfe",
        "gnomad41_exome_AF_fin", "gnomad41_exome_AF_asj", "gnomad41_exome_AF_remaining",
        "gnomad41_genome_AF", "gnomad41_genome_AF_afr", "gnomad41_genome_AF_amr",
        "gnomad41_genome_AF_asj", "gnomad41_genome_AF_eas", "gnomad41_genome_AF_fin",
        "gnomad41_genome_AF_nfe", "gnomad41_genome_AF_remaining",
        "HRC_AF", "HRC_AC", "HRC_AN", "HRC_non1000G_AF", "HRC_non1000G_AC", "HRC_non1000G_AN",
        "GME_AF", "GME_NWA", "GME_NEA", "GME_AP", "GME_Israel", "GME_SD", "GME_TP", "GME_CA",
        "cg69", "nci60",
    ],
}

# Display-label overrides for population frequency columns.
# Raw column names like AF_All are ambiguous — these prefixes clarify the source database.
POPULATION_FREQ_LABELS: dict = {
    "AF_All": "SG10K_AF_All",
    "AF_CHS": "SG10K_AF_CHS",
    "AF_INS": "SG10K_AF_INS",
    "AF_MAS": "SG10K_AF_MAS",
    "AF_SEA": "GenomeAsia_AF_SEA",
    "AF_NEA": "GenomeAsia_AF_NEA",
    "AF_SAS": "GenomeAsia_AF_SAS",
}

# ACMG criteria chip definitions: (header_name, label, css_class)
ACMG_CRITERIA = {
    "Very Strong Pathogenic": [("PVS1", "PVS1", "pvs")],
    "Strong Pathogenic": [("PS1", "PS1", "ps"), ("PS2", "PS2", "ps"), ("PS3", "PS3", "ps"), ("PS4", "PS4", "ps")],
    "Moderate Pathogenic": [("PM1", "PM1", "pm"), ("PM2", "PM2", "pm"), ("PM3", "PM3", "pm"),
                            ("PM4", "PM4", "pm"), ("PM5", "PM5", "pm"), ("PM6", "PM6", "pm")],
    "Supporting Pathogenic": [("PP1", "PP1", "pp"), ("PP2", "PP2", "pp"), ("PP3", "PP3", "pp"),
                               ("PP4", "PP4", "pp"), ("PP5", "PP5", "pp")],
    "Stand-Alone Benign": [("BA1", "BA1", "ba")],
    "Strong Benign": [("BS1", "BS1", "bs"), ("BS2", "BS2", "bs"), ("BS3", "BS3", "bs"), ("BS4", "BS4", "bs")],
    "Supporting Benign": [("BP1", "BP1", "bp"), ("BP2", "BP2", "bp"), ("BP3", "BP3", "bp"),
                          ("BP4", "BP4", "bp"), ("BP5", "BP5", "bp"), ("BP6", "BP6", "bp"), ("BP7", "BP7", "bp")],
}
# Flat lookup: criterion code → CSS modifier class
ACMG_CRITERION_CLASS = {
    criterion: css_class
    for criteria_list in ACMG_CRITERIA.values()
    for criterion, _label, css_class in criteria_list
}

# ClinVar column header names
CLINVAR_COL_NAMES = ["CLNALLELEID", "CLNDN", "CLNDISDB", "CLNREVSTAT", "CLNSIG"]

# Prediction score columns: (score_header, pred_header or None, display_label)
PREDICTION_SCORE_NAMES = [
    ("MCAP", None, "M-CAP"),
    ("REVEL", None, "REVEL"),
    ("SIFT_score", "SIFT_pred", "SIFT"),
    ("Polyphen2_HDIV_score", "Polyphen2_HDIV_pred", "PolyPhen2 HDIV"),
    ("Polyphen2_HVAR_score", "Polyphen2_HVAR_pred", "PolyPhen2 HVAR"),
    ("LRT_score", "LRT_pred", "LRT"),
    ("MutationTaster_score", "MutationTaster_pred", "MutationTaster"),
    ("MutationAssessor_score", "MutationAssessor_pred", "MutationAssessor"),
    ("FATHMM_score", "FATHMM_pred", "FATHMM"),
    ("MetaSVM_score", "MetaSVM_pred", "RadialSVM"),
    ("MetaLR_score", "MetaLR_pred", "LR"),
    ("VEST4_score", None, "VEST4"),
    ("CADD_raw", "CADD_phred", "CADD"),
    ("GERP++_RS", None, "GERP++"),
    ("phyloP30way_mammalian", None, "phyloP30way"),
    ("phyloP100way_vertebrate", None, "phyloP100way"),
    ("SiPhy_29way_logOdds", None, "SiPhy"),
    ("phastConsElements30way", None, "phastCons30way"),
    ("phastConsElements100way", None, "phastCons100way"),
    ("targetScanS", None, "targetScanS"),
]


def get_css_styles() -> str:
    return """
:root {
    --color-bg:         #f8fafc;
    --color-surface:    #ffffff;
    --color-surface-2:  #f1f5f9;
    --color-border:     #e2e8f0;
    --color-text:       #0f172a;
    --color-text-2:     #475569;
    --color-text-3:     #94a3b8;
    --color-accent:     #2563eb;
    --color-accent-bg:  #eff6ff;
    --color-path:       #dc2626;
    --color-lpath:      #ea580c;
    --color-benign:     #16a34a;
    --color-vus:        #64748b;
    /* legacy aliases kept for backward compat */
    --primary-color: #2c3e50;
    --secondary-color: #3498db;
    --accent-color: #e74c3c;
    --success-color: #27ae60;
    --warning-color: #f39c12;
    --light-bg: #f8f9fa;
    --border-color: #dee2e6;
    --text-muted: #6c757d;
    --tier1-color: #dc3545;
    --tier2-color: #fd7e14;
    --tier3-color: #ffc107;
    --tier4-color: #28a745;
    --pathogenic-color: #dc3545;
    --likely-pathogenic-color: #fd7e14;
    --vus-color: #6c757d;
    --likely-benign-color: #17a2b8;
    --benign-color: #28a745;
}

* { box-sizing: border-box; }

body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    line-height: 1.5;
    color: #212529;
    background-color: #f5f6fa;
    margin: 0;
    padding: 0;
}

.container {
    max-width: 1600px;
    margin: 0 auto;
    padding: 20px;
}

/* Header */
.header {
    background: #ffffff;
    border-bottom: 1px solid var(--color-border);
    border-left: 4px solid var(--color-accent);
    padding: 20px 32px;
    margin-bottom: 20px;
    border-radius: 8px;
}
.header h1 { margin: 0 0 4px 0; font-size: 1.75rem; font-weight: 700; color: var(--color-text); }
.header .subtitle { color: var(--color-text-2); font-size: 0.95rem; margin-top: 4px; }

/* Breadcrumb */
.breadcrumb {
    background: white;
    padding: 12px 20px;
    border-radius: 6px;
    margin-bottom: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
}
.breadcrumb a { color: var(--secondary-color); text-decoration: none; }
.breadcrumb a:hover { text-decoration: underline; }

/* Gene Hero */
.gene-hero {
    background: white;
    border-radius: 10px;
    padding: 25px 30px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 20px;
}
.gene-name { font-size: 2.8rem; font-weight: 700; color: var(--primary-color); margin: 0; }
.variant-notation { font-size: 1.4rem; color: var(--text-muted); font-family: 'Monaco','Menlo',monospace; }

/* Panel Styles */
.panel {
    background: white;
    border-radius: 10px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    overflow: hidden;
}
.panel-header {
    background: var(--light-bg);
    padding: 15px 20px;
    border-bottom: 1px solid var(--border-color);
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    transition: background-color 0.2s;
}
.panel-header:hover { background: #e9ecef; }
.panel-title { font-size: 1.1rem; font-weight: 600; color: var(--primary-color); margin: 0; }
.panel-toggle { font-size: 1.2rem; color: var(--text-muted); transition: transform 0.3s; }
.panel.collapsed .panel-toggle { transform: rotate(-90deg); }
.panel.collapsed .panel-content { display: none; }
.panel-content { padding: 20px; }

/* Data Grid */
.data-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 15px;
}
.data-item {
    padding: 10px 15px;
    background: var(--light-bg);
    border-radius: 6px;
    border-left: 3px solid var(--secondary-color);
}
.data-label {
    font-size: 0.75rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
}
.data-value { font-size: 0.875rem; color: #212529; word-break: break-word; }
.data-value.large { font-size: 1.2rem; font-weight: 600; }
.data-value.monospace { font-family: 'Monaco','Menlo',monospace; }

/* Population Frequency Bars */
.freq-bar-container {
    width: 100%; height: 6px; background: #e9ecef;
    border-radius: 3px; margin-top: 5px; overflow: hidden;
}
.freq-bar { height: 100%; border-radius: 3px; transition: width 0.3s; }
.freq-bar.rare { background: var(--success-color); }
.freq-bar.low { background: var(--warning-color); }
.freq-bar.common { background: var(--accent-color); }

/* ACMG Criteria Chips */
.acmg-grid { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 15px; }
.acmg-chip {
    display: inline-flex; align-items: center;
    padding: 6px 12px; border-radius: 20px;
    font-size: 0.8rem; font-weight: 600;
    background: #eff6ff; color: #1d4ed8;   /* default — overridden by type classes */
}
.acmg-chip.inactive { background: #e9ecef; color: #adb5bd; }
.acmg-chip.pvs { background: #721c24; color: white; }
.acmg-chip.ps  { background: var(--pathogenic-color); color: white; }
.acmg-chip.pm  { background: var(--likely-pathogenic-color); color: white; }
.acmg-chip.pp  { background: #f8d7da; color: #721c24; }
.acmg-chip.ba  { background: #155724; color: white; }
.acmg-chip.bs  { background: var(--benign-color); color: white; }
.acmg-chip.bp  { background: #d4edda; color: #155724; }

/* Prediction Score Table */
.score-table { width: 100%; border-collapse: collapse; font-size: 0.9rem; }
.score-table th, .score-table td {
    padding: 10px 12px; text-align: left;
    border-bottom: 1px solid var(--border-color);
}
.score-table th { background: var(--light-bg); font-weight: 600; color: var(--primary-color); }
.score-table tr:hover { background: #f8f9fa; }
.pred-badge {
    display: inline-block; padding: 3px 10px; border-radius: 12px;
    font-size: 0.75rem; font-weight: 600; text-transform: uppercase;
}
.pred-d, .pred-deleterious { background: #f8d7da; color: #721c24; }
.pred-t, .pred-tolerated { background: #d4edda; color: #155724; }
.pred-p, .pred-possibly { background: #fff3cd; color: #856404; }
.pred-b, .pred-benign { background: #d4edda; color: #155724; }
.pred-n, .pred-neutral { background: #e9ecef; color: #495057; }

/* Population DB Sub-panels */
.pop-db-section { margin-bottom: 20px; }
.pop-db-title {
    font-size: 0.9rem; font-weight: 600; color: var(--primary-color);
    margin-bottom: 10px; padding-bottom: 5px;
    border-bottom: 1px solid var(--border-color);
}

/* Panel Color Themes */
.panel-header {
    background: var(--color-surface-2);
    border-left: 3px solid var(--color-border);
    color: var(--color-text);
    padding: 10px 16px;
}
.panel.panel-basic      { border-left: 4px solid #2563eb; }
.panel.panel-basic      .panel-header { border-left-color: #2563eb; }
.panel.panel-population { border-left: 4px solid #7c3aed; }
.panel.panel-population .panel-header { border-left-color: #7c3aed; }
.panel.panel-acmg       { border-left: 4px solid #0891b2; }
.panel.panel-acmg       .panel-header { border-left-color: #0891b2; }
.panel.panel-computational { border-left: 4px solid #0d9488; }
.panel.panel-computational .panel-header { border-left-color: #0d9488; }
.panel.panel-comparison { border-left: 4px solid #16a34a; }
.panel.panel-comparison .panel-header { border-left-color: #16a34a; }
.panel.panel-additional { border-left: 4px solid #546e7a; }
.panel.panel-additional .panel-header { border-left-color: #546e7a; }
.panel.panel-igv        { border-left: 4px solid #8e44ad; }
.panel.panel-igv        .panel-header { border-left-color: #8e44ad; }
.panel.panel-igv .panel-content { text-align: center; }
.panel.panel-igv .igv-image {
    max-width: 100%; height: auto; border-radius: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.15); margin: 10px 0;
}
.panel.panel-igv .igv-no-image {
    color: var(--text-muted); font-style: italic; padding: 40px 20px;
}
.panel-title { font-size: 1.1rem; font-weight: 600; color: var(--color-text); margin: 0; }
.panel-toggle { font-size: 1.2rem; color: var(--color-text-3); transition: transform 0.3s; }

/* Dashboard Stats */
.dashboard-stats {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 15px;
    margin-bottom: 25px;
}
.stat-card {
    background: white; border-radius: 10px; padding: 20px;
    text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    transition: transform 0.2s, box-shadow 0.2s;
}
.stat-card.clickable { cursor: pointer; }
.stat-card.clickable:hover { transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,0.15); }
.stat-value { font-size: 2.5rem; font-weight: 700; color: var(--primary-color); }
.stat-label { font-size: 0.85rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.5px; }

/* CancerVar Tier Badge */
.cancervar-badge {
    display: inline-block; padding: 6px 16px; border-radius: 20px;
    font-size: 0.9rem; font-weight: 700; letter-spacing: 0.3px;
}
.cancervar-tier1 { background: var(--tier1-color); color: white; }
.cancervar-tier2 { background: var(--tier2-color); color: white; }
.cancervar-tier3 { background: var(--tier3-color); color: #212529; }
.cancervar-tier4 { background: var(--tier4-color); color: white; }
.cancervar-unknown { background: #6c757d; color: white; }

/* 0-variant sample cards */
.sample-card.no-variants .sample-card-header {
    background: linear-gradient(135deg, #6c757d 0%, #5a6268 100%);
}
.no-variants-badge {
    display: inline-block; padding: 2px 8px; border-radius: 10px;
    background: rgba(255,255,255,0.25); color: rgba(255,255,255,0.9);
    font-size: 0.75rem; margin-top: 5px;
}
.sample-stat-value.zero { color: var(--text-muted); }

/* VAF bar in variant table */
.vaf-bar-mini {
    width: 100%; height: 4px; background: #e9ecef;
    border-radius: 2px; margin-top: 4px; overflow: hidden; min-width: 50px;
}
.vaf-bar-mini-fill { height: 100%; border-radius: 2px; background: var(--secondary-color); }

/* Sample Cards */
.sample-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
    gap: 20px;
    margin-bottom: 30px;
}
.sample-card {
    background: white; border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    overflow: hidden; transition: transform 0.2s, box-shadow 0.2s;
    cursor: pointer; text-decoration: none; color: inherit; display: block;
}
.sample-card:hover { transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,0.15); }
.sample-card[data-tier="1"] { border-left: 4px solid #dc2626; }
.sample-card[data-tier="2"] { border-left: 4px solid #ea580c; }
.sample-card[data-tier="3"] { border-left: 4px solid var(--color-border); }
.sample-card-header {
    background: linear-gradient(135deg, var(--primary-color) 0%, #34495e 100%);
    color: white; padding: 20px;
}
.sample-card-header h3 { margin: 0; font-size: 1.2em; }
.sample-card-body { padding: 15px 20px; }
.sample-stat {
    display: flex; justify-content: space-between;
    padding: 8px 0; border-bottom: 1px solid var(--border-color);
}
.sample-stat:last-child { border-bottom: none; }
.sample-stat-label { color: var(--text-muted); font-size: 0.85rem; }
.sample-stat-value { font-weight: 600; }

/* Variant Table */
.variant-table {
    width: 100%; border-collapse: collapse; background: white;
    border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
.variant-table th, .variant-table td {
    padding: 14px 16px; text-align: left;
    border-bottom: 1px solid var(--border-color);
}
.variant-table th {
    background: var(--primary-color); color: white; font-weight: 600;
    font-size: 0.85rem; text-transform: uppercase; letter-spacing: 0.5px;
}
.variant-table tr:hover { background: #f8f9fa; }
.variant-table td.gene { font-weight: 600; color: var(--primary-color); }
.variant-table a { color: var(--secondary-color); text-decoration: none; }
.variant-table a:hover { text-decoration: underline; }

/* Back button */
.back-btn {
    display: inline-flex; align-items: center; gap: 8px;
    padding: 10px 18px; background: white;
    border: 1px solid var(--border-color); border-radius: 6px;
    color: var(--primary-color); text-decoration: none;
    font-size: 0.9rem; font-weight: 500; cursor: pointer; transition: all 0.2s;
}
.back-btn:hover { background: var(--secondary-color); color: white; border-color: var(--secondary-color); }

/* Responsive */
@media (max-width: 768px) {
    .container { padding: 10px; }
    .gene-hero { flex-direction: column; align-items: flex-start; }
    .gene-name { font-size: 2rem; }
    .data-grid { grid-template-columns: 1fr; }
    .sample-grid { grid-template-columns: 1fr; }
}

@media print {
    .panel.collapsed .panel-content { display: block !important; }
    .breadcrumb { display: none; }
    .panel { break-inside: avoid; }
}

/* ── Tab bar ─────────────────────────────────────────────────────── */
.tab-bar {
    display: flex;
    gap: 0;
    margin: 0 0 16px 0;
    border-bottom: 2px solid var(--color-border);
    padding-bottom: 0;
}
.tab-btn {
    background: none;
    border: none;
    border-bottom: 3px solid transparent;
    margin-bottom: -2px;
    padding: 10px 20px;
    cursor: pointer;
    font-size: 0.95em;
    color: var(--color-text-2);
    font-weight: 500;
    position: relative;
}
.tab-btn:hover { color: var(--color-text); background: var(--color-surface-2); }
.tab-btn.active {
    color: var(--color-accent);
    border-bottom-color: var(--color-accent);
    background: var(--color-accent-bg);
    font-weight: 600;
}

/* ── Clinical summary table ───────────────────────────────────────── */
.clinical-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.9em;
    margin-bottom: 24px;
}
.clinical-table th {
    background: var(--color-surface-2);
    color: var(--color-text-3);
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    padding: 10px 12px;
    text-align: left;
    border-bottom: 2px solid var(--color-border);
    white-space: nowrap;
}
.clinical-table td {
    padding: 10px 12px;
    border-bottom: 1px solid #e9ecef;
    vertical-align: middle;
}
.clinical-table tbody tr.summary-row:nth-child(4n+1) { background: #fafbfc; }
.clinical-table td.col-hgvsc { font-family: 'Monaco','Menlo',monospace; }
.clinical-table tr:hover { background: #f8f9fa; cursor: pointer; }

/* ── ClinVar / InterVar badges ────────────────────────────────────── */
.clnsig-badge, .intervar-badge, .cancervar-badge {
    display: inline-flex;
    align-items: center;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.78rem;
    font-weight: 600;
    white-space: nowrap;
    letter-spacing: 0.2px;
}
.badge-pathogenic        { background: #fee2e2; color: #b91c1c; }
.badge-likely-pathogenic { background: #ffedd5; color: #c2410c; }
.badge-benign            { background: #dcfce7; color: #15803d; }
.badge-likely-benign     { background: #d1fae5; color: #065f46; }
.badge-vus               { background: #f1f5f9; color: var(--color-text-2); }
.badge-other             { background: #f1f5f9; color: var(--color-text-2); }

/* ── CLNREVSTAT star rating ───────────────────────────────────────── */
.clnrev-stars {
    font-size: 1em;
    letter-spacing: 1px;
    color: #f5a623;
    cursor: help;
}

/* ── Variant detail cards ─────────────────────────────────────────── */
.variant-card {
    border: 1px solid #dee2e6;
    border-radius: 6px;
    margin-bottom: 12px;
    overflow: hidden;
}
.card-header {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 8px;
    padding: 10px 14px;
    background: #f8f9fa;
    cursor: pointer;
    user-select: none;
}
.card-header:hover { background: #e9ecef; }
.card-gene {
    font-weight: 700;
    font-size: 1.05em;
    margin-right: 4px;
}
.card-hgvs {
    color: #495057;
    font-family: monospace;
    font-size: 0.9em;
}
.card-vaf {
    margin-left: auto;
    font-size: 0.88em;
    color: #6c757d;
}
.card-toggle {
    font-size: 0.8em;
    color: #6c757d;
    margin-left: 6px;
}
.card-body {
    padding: 12px 16px;
    font-size: 0.9em;
    border-top: 1px solid #dee2e6;
}
.card-body.collapsed { display: none; }
.card-row {
    display: flex;
    gap: 8px;
    margin-bottom: 6px;
    align-items: flex-start;
}
.card-label {
    color: var(--color-text-2);
    font-size: 0.8rem;
    text-transform: uppercase;
    letter-spacing: 0.4px;
    min-width: 140px;
    flex-shrink: 0;
}
/* Size-only overrides for compact contexts — colours come from type classes above */
.card-body .acmg-chip,
.detail-stack .acmg-chip {
    padding: 1px 7px; border-radius: 12px;
    font-size: 0.75rem; display: inline-block; margin: 1px 2px;
}
/* Stat pill for single values (gnomAD etc.) */
.stat-pill {
    display: inline-block; background: var(--color-surface-2);
    border: 1px solid var(--color-border); color: var(--color-text);
    padding: 1px 9px; border-radius: 12px; font-size: 0.82rem; font-family: 'Monaco','Menlo',monospace;
}
/* Review status / DB pills */
.revstat-pill {
    background: #f0fdf4; color: #166534; border: 1px solid #bbf7d0;
    padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; display: inline-block;
}
.disdb-pill {
    background: #f8fafc; color: var(--color-text-2); border: 1px solid var(--color-border);
    padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; font-family: 'Monaco','Menlo',monospace; display: inline-block;
}
.card-igv img {
    max-width: 100%;
    border: 1px solid #dee2e6;
    border-radius: 4px;
    margin-top: 6px;
}

/* ── Expandable detail rows (clinical table) ──────────────────────── */
.clinical-table .toggle-btn {
    background: none;
    border: none;
    cursor: pointer;
    font-size: 0.8em;
    color: #6c757d;
    padding: 0 4px;
}
.detail-row.collapsed { display: none; }
.detail-row > td {
    padding: 0;
    border-bottom: 2px solid var(--color-border);
}
.detail-content {
    background: #f8fafc;
    border-left: 3px solid var(--color-accent);
    padding: 16px 20px;
    display: flex;
    flex-wrap: wrap;
    gap: 20px;
    align-items: flex-start;
}
.detail-top {
    display: flex; flex-wrap: wrap; gap: 20px; align-items: center;
    margin-bottom: 12px; padding-bottom: 12px;
    border-bottom: 1px solid var(--color-border);
}
.detail-top .card-row { margin-bottom: 0; }
.detail-bottom { display: flex; gap: 20px; align-items: flex-start; }
.detail-associations { flex: 1; min-width: 0; }
.detail-stack { flex: 1; min-width: 0; display: flex; flex-direction: column; gap: 10px; }
.detail-stack-row { display: flex; flex-direction: column; gap: 4px; }
.detail-stack-row .card-label { margin-bottom: 2px; }
.detail-igv { flex: 0 0 auto; }
.detail-igv img {
    max-width: 500px; width: 100%;
    border: 1px solid #dee2e6; border-radius: 4px;
}
/* Disease pills */
.disease-section { margin-bottom: 10px; }
.disease-section-label {
    font-size: 0.72rem; font-weight: 600; text-transform: uppercase;
    letter-spacing: 0.5px; color: var(--color-text-3); margin-bottom: 6px;
}
.disease-pills { display: flex; flex-wrap: wrap; gap: 4px; align-items: center; }
.disease-pill {
    background: #f0f4ff; color: #3b4a6b; border: 1px solid #c7d2fe;
    padding: 2px 8px; border-radius: 12px; font-size: 0.78rem;
}
.disease-more-btn, .cosmic-more-btn {
    background: none; border: 1px solid var(--color-border);
    color: var(--color-accent); border-radius: 12px;
    padding: 2px 8px; font-size: 0.75rem; cursor: pointer; font-weight: 600;
}
.disease-more.hidden, .cosmic-more.hidden { display: none; }
/* COSMIC tissue chips */
.cosmic-section { margin-top: 8px; }
.cosmic-id { font-family: monospace; font-size: 0.8rem; color: var(--color-text-2); }
.cosmic-tissues { display: flex; flex-wrap: wrap; gap: 4px; margin-top: 6px; align-items: center; }
.cosmic-tissue {
    background: #fef3c7; color: #92400e; border: 1px solid #fde68a;
    padding: 2px 8px; border-radius: 12px; font-size: 0.75rem;
}
.cosmic-tissue strong { font-weight: 700; }
"""


def get_javascript() -> str:
    return """
function togglePanel(header) {
    header.parentElement.classList.toggle('collapsed');
}

function switchTab(btn, tabId) {
    // Deactivate all tabs in the same tab-bar
    var bar = btn.closest('.tab-bar');
    bar.querySelectorAll('.tab-btn').forEach(function(b) {
        b.classList.remove('active');
    });
    // Hide all sibling tab-content divs
    var container = bar.parentElement;
    container.querySelectorAll('.tab-content').forEach(function(div) {
        div.style.display = 'none';
    });
    // Activate clicked tab and show its content
    btn.classList.add('active');
    var panel = container.querySelector('#' + tabId);
    if (panel) panel.style.display = '';
}

function toggleCard(header) {
    var card = header.closest('.variant-card');
    var body = card.querySelector('.card-body');
    var toggle = header.querySelector('.card-toggle');
    body.classList.toggle('collapsed');
    if (toggle) {
        toggle.textContent = body.classList.contains('collapsed') ? '▶' : '▼';
    }
}

function scrollToCard(cardId) {
    var el = document.getElementById(cardId);
    if (el) el.scrollIntoView({ behavior: 'smooth', block: 'start' });
}

function toggleDetailRow(btn) {
    var detailRow = btn.closest('tr').nextElementSibling;
    detailRow.classList.toggle('collapsed');
    btn.textContent = detailRow.classList.contains('collapsed') ? '▶' : '▼';
}

function toggleMore(btn, targetId) {
    var el = document.getElementById(targetId);
    if (!el) return;
    var hidden = el.classList.toggle('hidden');
    btn.textContent = hidden ? btn.dataset.more : btn.dataset.less;
}
"""


def _is_empty(val):
    """Check if value is empty/missing"""
    if val is None:
        return True
    s = str(val).strip()
    return s in ("", ".", "-", "NA", "N/A", "None")


def _pred_badge_class(pred_str):
    """Get CSS class for prediction badge"""
    if not pred_str:
        return ""
    p = pred_str.strip().upper()[0] if pred_str.strip() else ""
    mapping = {"D": "pred-d", "T": "pred-t", "P": "pred-p", "B": "pred-b", "N": "pred-n"}
    return mapping.get(p, "")


def _cancervar_tier_badge(cancervar_str: str) -> str:
    """Generate CancerVar tier badge HTML from 'CancerVar and Evidence' column value."""
    if not cancervar_str:
        return ""
    # Check IV before III before II before I to avoid shorter patterns matching first
    tier_map = [
        (r'Tier[_ ]?IV',  "tier4", "Tier IV &mdash; Benign / Likely Benign"),
        (r'Tier[_ ]?III', "tier3", "Tier III &mdash; Unknown Significance"),
        (r'Tier[_ ]?II',  "tier2", "Tier II &mdash; Potentially Actionable"),
        (r'Tier[_ ]?I',   "tier1", "Tier I &mdash; Actionable"),
    ]
    for pattern, css_cls, label in tier_map:
        if re.search(pattern, cancervar_str, re.IGNORECASE):
            return f'<span class="cancervar-badge cancervar-{css_cls}">{label}</span>'
    return ""


_CLNREVSTAT_STARS = {
    "practice_guideline": "★★★★",
    "reviewed_by_expert_panel": "★★★",
    "criteria_provided,_multiple_submitters,_no_conflicts": "★★",
    "criteria_provided,_single_submitter": "★",
    "no_assertion_criteria_provided": "☆",
    "no_assertion_provided": "☆",
    "conflicting_interpretations_of_pathogenicity": "⚠ conflict",
}


def clnrevstat_to_stars(val: str) -> str:
    """Convert CLNREVSTAT string to star HTML (span with title)."""
    val = (val or "").strip()
    stars = _CLNREVSTAT_STARS.get(val, "")
    if not stars:
        return ""
    return f'<span class="clnrev-stars" title="{escape(val)}">{stars}</span>'


def _clinvar_sig_badge(sig: str) -> str:
    """Return a coloured ClinVar significance badge.
    Expected input: ANNOVAR ClinVar format with underscores (e.g. 'Likely_pathogenic').
    """
    sig = (sig or "").strip()
    if not sig or sig in (".", "not_provided"):
        return ""
    sig_lower = sig.lower().replace(" ", "_")
    if "pathogenic" in sig_lower and "likely" not in sig_lower and "benign" not in sig_lower:
        css = "badge-pathogenic"
    elif "likely_pathogenic" in sig_lower:
        css = "badge-likely-pathogenic"
    elif "benign" in sig_lower and "likely" not in sig_lower:
        css = "badge-benign"
    elif "likely_benign" in sig_lower:
        css = "badge-likely-benign"
    elif "uncertain" in sig_lower or "vus" in sig_lower:
        css = "badge-vus"
    else:
        css = "badge-other"
    return f'<span class="clnsig-badge {css}">{escape(sig)}</span>'


def _intervar_badge(intervar: str) -> str:
    """Return a coloured InterVar classification badge.
    Expected input: InterVar format with spaces (e.g. 'Likely pathogenic').
    """
    intervar = (intervar or "").strip()
    if not intervar or intervar in (".", "Unknown"):
        return ""
    iv_lower = intervar.lower()
    if "pathogenic" in iv_lower and "likely" not in iv_lower:
        css = "badge-pathogenic"
    elif "likely pathogenic" in iv_lower:
        css = "badge-likely-pathogenic"
    elif "benign" in iv_lower and "likely" not in iv_lower:
        css = "badge-benign"
    elif "likely benign" in iv_lower:
        css = "badge-likely-benign"
    elif "uncertain" in iv_lower:
        css = "badge-vus"
    else:
        css = "badge-other"
    return f'<span class="intervar-badge {css}">{escape(intervar)}</span>'


class iSeqReportGenerator:
    def __init__(self, excel_path: str, output_dir: str, snapshots_dir: Optional[str] = None):
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.snapshots_dir = Path(snapshots_dir) if snapshots_dir else None
        self.headers: List[str] = []
        self.rows: List[List[Any]] = []
        self.samples: Dict[str, List[int]] = defaultdict(list)
        self.col_idx: Dict[str, int] = {}

    def load_excel(self):
        print(f"Loading Excel file: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        self.headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(header_row)]
        print(f"Found {len(self.headers)} columns")

        # Detect per-sample xlsx (no SAMPLE column): col 1 is Chr, not SAMPLE.
        # Prepend a synthetic SAMPLE column so all column name lookups stay correct.
        is_per_sample = self.headers[0].upper() != "SAMPLE"
        if is_per_sample:
            inferred_sample = self.excel_path.stem
            self.headers = ["SAMPLE"] + self.headers
            print(f"Per-sample format detected, using sample name: {inferred_sample}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row_data = list(row)
            if is_per_sample:
                row_data = [inferred_sample] + row_data
            while len(row_data) < len(self.headers):
                row_data.append(None)
            self.rows.append(row_data)
            sample_name = str(row_data[0]).strip() if row_data[0] else "Unknown"
            self.samples[sample_name].append(row_idx)

        wb.close()
        print(f"Loaded {len(self.rows)} variants from {len(self.samples)} samples")

        # Build header-name → 1-based column index mapping (strip whitespace for robustness)
        self.col_idx: Dict[str, int] = {}
        for i, h in enumerate(self.headers, 1):
            if h and not h.startswith("Column_"):
                self.col_idx[h] = i
                self.col_idx[h.strip()] = i

    def _col(self, name: str) -> Optional[int]:
        """Return 1-based column index by header name, or None if not found."""
        return self.col_idx.get(name) or self.col_idx.get(name.strip())

    def _cols(self, names: List[str]) -> List[int]:
        """Return list of 1-based column indices for the given header names, skipping missing."""
        result = []
        for name in names:
            idx = self._col(name)
            if idx:
                result.append(idx)
        return result

    def _val_n(self, row_idx: int, name: str) -> str:
        """Get cell value by column header name."""
        col = self._col(name)
        return self._val(row_idx, col) if col else ""

    def _val(self, row_idx: int, col_idx: int) -> str:
        """Get cell value (1-indexed col), returns '' if empty"""
        if col_idx < 1 or col_idx > len(self.headers) or row_idx < 0 or row_idx >= len(self.rows):
            return ""
        val = self.rows[row_idx][col_idx - 1]
        if _is_empty(val):
            return ""
        return str(val).strip()

    def _header(self, col_idx: int) -> str:
        if col_idx < 1 or col_idx > len(self.headers):
            return f"Column_{col_idx}"
        return self.headers[col_idx - 1]

    def _get_active_acmg_criteria(self, row, col_map):
        """Return list of ACMG criterion names that have evidence in this row.

        A criterion is active if its column value is non-empty, not '.', and not '0'.
        Iterates all criteria defined in ACMG_CRITERIA (e.g. PVS1, PS1, PM2, etc.).
        """
        active = []
        for _category, criteria_list in ACMG_CRITERIA.items():
            for criterion, _label, _css_class in criteria_list:
                col_idx = col_map.get(criterion)
                if col_idx is None:
                    continue
                val = str(row[col_idx]).strip() if col_idx < len(row) else ""
                if val and val not in (".", "0"):
                    active.append(criterion)
        return active

    @staticmethod
    def _clndn_pills_html(clndn_raw: str, uid: str, max_visible: int = 8) -> str:
        """Render ClinVar disease names as flex-wrap pill chips with +N more toggle."""
        from html import escape as _esc
        diseases = [d.replace('_', ' ').strip() for d in clndn_raw.split('|')
                    if d.strip() and d.strip() not in ('.', '')]
        if not diseases:
            return ''
        visible = diseases[:max_visible]
        hidden = diseases[max_visible:]
        pills = ''.join(f'<span class="disease-pill">{_esc(d)}</span>' for d in visible)
        if hidden:
            more_pills = ''.join(f'<span class="disease-pill">{_esc(d)}</span>' for d in hidden)
            pills += (f'<span class="disease-more hidden" id="dm-{uid}">{more_pills}</span>'
                      f'<button class="disease-more-btn" data-more="+{len(hidden)} more"'
                      f' data-less="show less" onclick="toggleMore(this,\'dm-{uid}\')">'
                      f'+{len(hidden)} more</button>')
        return (f'<div class="disease-section">'
                f'<div class="disease-section-label">ClinVar diseases</div>'
                f'<div class="disease-pills">{pills}</div></div>')

    @staticmethod
    def _cosmic_html(cosmic_raw: str, uid: str, max_visible: int = 5, show_label: bool = True) -> str:
        """Render COSMIC ID + top tissue occurrences as chips with +N more toggle."""
        import re as _re
        from html import escape as _esc
        parts = cosmic_raw.split(';')
        cosmic_id = parts[0].strip()
        occ_str = next((p for p in parts if p.startswith('OCCURENCE=')), '')
        if not occ_str:
            header = (f'<div class="disease-section-label">COSMIC '
                      f'<span class="cosmic-id">{_esc(cosmic_id)}</span></div>') if show_label else ''
            return f'<div class="cosmic-section">{header}<span class="cosmic-id">{_esc(cosmic_id)}</span></div>'
        tissues = []
        for item in occ_str[len('OCCURENCE='):].split(','):
            m = _re.match(r'(\d+)\(([^)]+)\)', item.strip())
            if m:
                tissues.append((int(m.group(1)), m.group(2).replace('_', ' ')))
        tissues.sort(reverse=True)

        def _chip(count, tissue):
            return f'<span class="cosmic-tissue">{_esc(tissue)} <strong>{count}</strong></span>'

        visible = tissues[:max_visible]
        hidden = tissues[max_visible:]
        chips = ''.join(_chip(c, t) for c, t in visible)
        if hidden:
            more_chips = ''.join(_chip(c, t) for c, t in hidden)
            chips += (f'<span class="cosmic-more hidden" id="cm-{uid}">{more_chips}</span>'
                      f'<button class="cosmic-more-btn" data-more="+{len(hidden)} tissues"'
                      f' data-less="show less" onclick="toggleMore(this,\'cm-{uid}\')">'
                      f'+{len(hidden)} tissues</button>')
        header = (f'<div class="disease-section-label">COSMIC '
                  f'<span class="cosmic-id">{_esc(cosmic_id)}</span></div>') if show_label else ''
        return f'<div class="cosmic-section">{header}<div class="cosmic-tissues">{chips}</div></div>'

    def _clinical_summary_table_html(self, variants, col_map, sample_html_dir=None):
        """Build expandable clinical summary table — one summary row + one detail row per variant.

        sample_html_dir: Path or str of the directory where the sample HTML file will be saved,
                         used to compute correct relative paths for IGV screenshots.
        """
        from html import escape
        import os as _os

        def _get(row, col_name, default=""):
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row):
                return default
            val = str(row[idx]).strip()
            return val if val not in (".", "") else default

        rows_html = []
        for i, row in enumerate(variants):
            gene   = escape(_get(row, "GENE"))
            hgvsc  = escape(_get(row, "HGVSc"))
            hgvsp  = escape(_get(row, "HGVSp"))
            clnsig = _get(row, "CLNSIG")
            clnrev = _get(row, "CLNREVSTAT")
            canvar = _get(row, "CancerVar and Evidence")
            interv = _get(row, "InterVar_automated")
            vaf    = escape(_get(row, "VAF"))
            dp     = escape(_get(row, "DP"))
            clndn_raw = _get(row, "CLNDN")
            clndn  = escape(clndn_raw)
            gnomad = escape(_get(row, "gnomad41_genome_AF"))
            cosmic_raw = _get(row, "cosmic91")
            cosmic = escape(cosmic_raw)

            clnsig_badge = _clinvar_sig_badge(clnsig)   if clnsig  else ""
            stars        = clnrevstat_to_stars(clnrev)   if clnrev  else ""
            canvar_badge = _cancervar_tier_badge(canvar) if canvar  else ""
            interv_badge = _intervar_badge(interv)       if interv  else ""

            active_criteria = self._get_active_acmg_criteria(row, col_map)
            acmg_chips = " ".join(
                f'<span class="acmg-chip {ACMG_CRITERION_CLASS.get(c, "")}">{escape(c)}</span>'
                for c in active_criteria
            )

            # IGV screenshot — compute path relative to the sample HTML file directory
            igv_html = ""
            sample_val = _get(row, "SAMPLE")
            chrom_val  = _get(row, "Chr")
            pos_val    = _get(row, "Pos")
            if sample_val and chrom_val and pos_val:
                igv_path = self.find_igv_screenshot(sample_val, chrom_val, pos_val)
                if igv_path and sample_html_dir:
                    rel_igv = _os.path.relpath(igv_path, str(sample_html_dir))
                    igv_html = f'<div class="detail-igv"><img src="{escape(rel_igv)}" alt="IGV snapshot"></div>'

            # Build stacked detail rows (left column), IGV on the right
            uid = f"{i}"
            stack_rows = []
            if gnomad:
                stack_rows.append(
                    f'<div class="detail-stack-row">'
                    f'<span class="card-label">gnomAD AF</span>'
                    f'<span><span class="stat-pill">{gnomad}</span></span>'
                    f'</div>')
            if active_criteria:
                stack_rows.append(
                    f'<div class="detail-stack-row">'
                    f'<span class="card-label">ACMG criteria</span>'
                    f'<span>{acmg_chips}</span>'
                    f'</div>')
            if clndn_raw:
                stack_rows.append(
                    f'<div class="detail-stack-row">'
                    f'<span class="card-label">ClinVar diseases</span>'
                    f'{self._clndn_pills_html(clndn_raw, uid)}'
                    f'</div>')
            if cosmic_raw:
                stack_rows.append(
                    f'<div class="detail-stack-row">'
                    f'{self._cosmic_html(cosmic_raw, uid)}'
                    f'</div>')
            stack_html = "\n".join(stack_rows)

            # Summary row (always visible) + detail row (toggle)
            rows_html.append(f"""
        <tr class="summary-row" onclick="scrollToCard('card-{i}')">
            <td><button class="toggle-btn" onclick="toggleDetailRow(this)">▶</button></td>
            <td>{gene}</td>
            <td class="col-hgvsc">{hgvsc}</td>
            <td>{hgvsp}</td>
            <td>{clnsig_badge or escape(clnsig)}</td>
            <td>{stars}</td>
            <td>{canvar_badge}</td>
            <td>{interv_badge or escape(interv)}</td>
            <td>{vaf}</td>
            <td>{dp}</td>
        </tr>
        <tr class="detail-row collapsed">
            <td colspan="10">
              <div class="detail-content">
                <div class="detail-bottom">
                  <div class="detail-stack">{stack_html}</div>
                  {igv_html}
                </div>
              </div>
            </td>
        </tr>""")

        rows_str = "\n".join(rows_html)
        return f"""
<table class="clinical-table">
  <thead>
    <tr>
      <th style="width:2em"></th>
      <th>Gene</th>
      <th>HGVSc</th>
      <th>HGVSp</th>
      <th>ClinVar</th>
      <th>&#9733;</th>
      <th>CancerVar</th>
      <th>InterVar</th>
      <th>VAF</th>
      <th>DP</th>
    </tr>
  </thead>
  <tbody>
    {rows_str}
  </tbody>
</table>"""

    def _variant_detail_card_html(self, i, row, col_map, snapshot_dir, collapsed=False):
        """Build HTML for a single collapsible variant detail card.

        i: 0-based variant index (used for card id and scrollToCard anchor)
        row: data row (list of cell values)
        col_map: dict mapping column header strings to column indices
        snapshot_dir: path to IGV snapshot directory (string or Path); unused
                      (IGV lookup uses self.snapshots_dir via find_igv_screenshot)
        collapsed: if True, card-body starts collapsed
        """
        from html import escape

        def _get(col_name, default=""):
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row):
                return default
            val = str(row[idx]).strip()
            return val if val not in (".", "") else default

        gene   = escape(_get("GENE"))
        hgvsc  = escape(_get("HGVSc"))
        hgvsp  = escape(_get("HGVSp"))
        clnsig = _get("CLNSIG")
        clnrev = _get("CLNREVSTAT")
        canvar = _get("CancerVar and Evidence")
        interv = _get("InterVar_automated")
        vaf    = escape(_get("VAF"))
        dp     = escape(_get("DP"))
        clndn  = escape(_get("CLNDN"))

        gnomad  = escape(_get("gnomad41_genome_AF"))
        cosmic = escape(_get("cosmic91"))

        clnsig_badge = _clinvar_sig_badge(clnsig)    if clnsig  else ""
        stars        = clnrevstat_to_stars(clnrev)    if clnrev  else ""
        canvar_badge = _cancervar_tier_badge(canvar)  if canvar  else ""
        interv_badge = _intervar_badge(interv)         if interv  else ""

        # Active ACMG criteria chips
        active_criteria = self._get_active_acmg_criteria(row, col_map)
        acmg_chips = " ".join(
            f'<span class="acmg-chip">{escape(c)}</span>'
            for c in active_criteria
        )

        # IGV screenshot — find_igv_screenshot needs sample, chrom, pos
        igv_html = ""
        sample_val = _get("SAMPLE")
        chrom_val  = _get("Chr")
        pos_val    = _get("Pos")
        if sample_val and chrom_val and pos_val:
            igv_path = self.find_igv_screenshot(sample_val, chrom_val, pos_val)
            if igv_path:
                rel = Path(igv_path).name
                igv_html = f'<div class="card-igv"><img src="SnapShots/{escape(rel)}" alt="IGV snapshot"></div>'

        collapsed_class = " collapsed" if collapsed else ""
        toggle_symbol = "▶" if collapsed else "▼"

        return f"""
<div class="variant-card" id="card-{i}">
  <div class="card-header" onclick="toggleCard(this)">
    <span class="card-gene">{gene}</span>
    <span class="card-hgvs">{hgvsc}</span>
    {f'<span class="card-hgvs">· {hgvsp}</span>' if hgvsp else ''}
    {clnsig_badge}
    {stars}
    {canvar_badge}
    <span class="card-vaf">VAF: {vaf} · DP: {dp}</span>
    <span class="card-toggle">{toggle_symbol}</span>
  </div>
  <div class="card-body{collapsed_class}">
    {f'<div class="card-row"><span class="card-label">ClinVar disease:</span><span>{clndn}</span></div>' if clndn else ''}
    {f'<div class="card-row"><span class="card-label">ACMG criteria:</span><span>{acmg_chips}</span></div>' if active_criteria else ''}
    {f'<div class="card-row"><span class="card-label">gnomAD AF:</span><span>{gnomad}</span></div>' if gnomad else ''}
    {f'<div class="card-row"><span class="card-label">COSMIC ID:</span><span>{cosmic}</span></div>' if cosmic else ''}
    {igv_html}
  </div>
</div>"""

    def find_igv_screenshot(self, sample: str, chrom: str, pos: str) -> Optional[str]:
        if not self.snapshots_dir or not self.snapshots_dir.exists():
            return None
        sample = sample.strip()
        chrom = chrom.strip()
        pos = pos.strip()
        chr_num = chrom.replace("chr", "")

        for filename in [f"{sample}-{chr_num}-{pos}.png", f"{sample}-chr{chr_num}-{pos}.png",
                         f"{sample}_{chr_num}_{pos}.png", f"{sample}_{chrom}_{pos}.png"]:
            filepath = self.snapshots_dir / filename
            if filepath.exists():
                return str(filepath.resolve())
        return None

    def _format_data_value(self, col_header: str, value: str, uid: str = "") -> str:
        """Format a cell value as HTML, adding hyperlinks for known column types."""
        if not value:
            return escape(value)

        # CLNDN — pipe-separated disease names → pills, max 6 visible
        if col_header == "CLNDN":
            diseases = [d.replace('_', ' ').strip() for d in value.split('|')
                        if d.strip() and d.strip() not in ('.', '')]
            if diseases:
                visible, hidden = diseases[:6], diseases[6:]
                pills = ''.join(f'<span class="disease-pill">{escape(d)}</span>' for d in visible)
                if hidden:
                    more_id = f"dn-{uid}"
                    pills += (f'<span class="disease-more hidden" id="{more_id}">'
                              + ''.join(f'<span class="disease-pill">{escape(d)}</span>' for d in hidden)
                              + f'</span><button class="disease-more-btn" data-more="+{len(hidden)} more"'
                              f' data-less="show less" onclick="toggleMore(this,\'{more_id}\')">'
                              f'+{len(hidden)} more</button>')
                return f'<div class="disease-pills" style="margin-top:4px">{pills}</div>'

        # CLNDISDB — pipe-separated DB:ID entries → monospace pills, max 5 visible
        if col_header == "CLNDISDB":
            entries = [e.strip() for e in value.split('|') if e.strip() and e.strip() != '.']
            if entries:
                visible, hidden = entries[:5], entries[5:]
                pills = ''.join(f'<span class="disdb-pill">{escape(e)}</span>' for e in visible)
                if hidden:
                    more_id = f"db-{uid}"
                    pills += (f'<span class="disease-more hidden" id="{more_id}">'
                              + ''.join(f'<span class="disdb-pill">{escape(e)}</span>' for e in hidden)
                              + f'</span><button class="disease-more-btn" data-more="+{len(hidden)} more"'
                              f' data-less="show less" onclick="toggleMore(this,\'{more_id}\')">'
                              f'+{len(hidden)} more</button>')
                return f'<div style="display:flex;flex-wrap:wrap;gap:3px;margin-top:4px">{pills}</div>'

        # CLNREVSTAT — pipe-separated review statuses → green pills
        if col_header == "CLNREVSTAT":
            statuses = [s.replace('_', ' ').strip() for s in value.split('|')
                        if s.strip() and s.strip() != '.']
            if statuses:
                pills = ''.join(f'<span class="revstat-pill">{escape(s)}</span> ' for s in statuses)
                return f'<div style="display:flex;flex-wrap:wrap;gap:3px;margin-top:4px">{pills}</div>'

        # COSMIC91 — tissue occurrence chips (no label; data-label row provides it)
        if col_header == "cosmic91":
            return self._cosmic_html(value, uid or "fa", show_label=False)

        return escape(value)

    def _data_grid_html(self, row_idx: int, col_indices: List[int]) -> str:
        """Generate data grid HTML, skipping empty values"""
        items = []
        for col_idx in col_indices:
            value = self._val(row_idx, col_idx)
            if not value:
                continue
            col_header = self._header(col_idx)
            label = escape(col_header)
            uid = f"fa-{row_idx}-{col_idx}"
            value_html = self._format_data_value(col_header, value, uid)
            items.append(f'''
                    <div class="data-item">
                        <div class="data-label">{label}</div>
                        <div class="data-value monospace">{value_html}</div>
                    </div>''')
        if not items:
            return '<div style="padding:10px;color:var(--text-muted);font-style:italic;">No data available</div>'
        return '<div class="data-grid">' + ''.join(items) + '\n</div>'

    def _panel_html(self, title: str, panel_class: str, content: str, collapsed: bool = False) -> str:
        cls = f"panel {panel_class}" + (" collapsed" if collapsed else "")
        return f'''
        <div class="{cls}">
            <div class="panel-header" onclick="togglePanel(this)">
                <h3 class="panel-title">{escape(title)}</h3>
                <span class="panel-toggle">&#9662;</span>
            </div>
            <div class="panel-content">
                {content}
            </div>
        </div>'''

    def _acmg_section_html(self, row_idx: int) -> str:
        """Generate ACMG section with chip buttons"""
        html = ''

        # InterVar classification
        intervar = self._val_n(row_idx, "InterVar_automated")
        if intervar:
            html += f'''
                <div style="margin-bottom: 20px;">
                    <div class="data-label">InterVar Classification</div>
                    <div class="data-value large">{escape(intervar)}</div>
                </div>'''

        # ACMG Criteria chips
        html += '<div style="margin-bottom: 25px;">\n'
        html += '    <div class="data-label" style="margin-bottom: 10px;">ACMG Criteria</div>\n'

        for group_name, criteria in ACMG_CRITERIA.items():
            html += f'    <div style="margin-bottom: 8px;">\n'
            html += f'        <small style="color: var(--text-muted);">{escape(group_name)}</small>\n'
            html += '        <div class="acmg-grid">\n'

            for col_name, label, css_class in criteria:
                value = self._val_n(row_idx, col_name)
                is_active = bool(value) and value != "0"
                chip_class = css_class if is_active else "inactive"
                html += f'            <span class="acmg-chip {chip_class}">{label}</span>\n'

            html += '        </div>\n'
            html += '    </div>\n'

        html += '</div>\n'

        # ClinVar data
        clinvar_items = []
        for col_idx in self._cols(CLINVAR_COL_NAMES):
            value = self._val(row_idx, col_idx)
            if value:
                col_header = self._header(col_idx)
                label = escape(col_header)
                # Link CLNALLELEID values to ClinVar
                if col_header == "CLNALLELEID":
                    parts = re.split(r'[;,]', value)
                    linked = []
                    for part in parts:
                        part = part.strip()
                        if part.isdigit():
                            linked.append(
                                f'<a href="https://www.ncbi.nlm.nih.gov/clinvar/variation/{part}/"'
                                f' target="_blank">{escape(part)}</a>'
                            )
                        elif part:
                            linked.append(escape(part))
                    value_html = '; '.join(linked) if linked else escape(value)
                else:
                    value_html = self._format_data_value(
                        col_header, value, uid=f"acmg-{row_idx}-{col_idx}")
                clinvar_items.append(f'''
                    <div class="data-item">
                        <div class="data-label">{label}</div>
                        <div class="data-value monospace">{value_html}</div>
                    </div>''')

        if clinvar_items:
            html += '<div class="data-label" style="margin-bottom: 10px;">ClinVar</div>\n'
            html += '<div class="data-grid">' + ''.join(clinvar_items) + '</div>\n'

        return html

    def _computational_section_html(self, row_idx: int) -> str:
        """Generate computational predictions as a score table"""
        rows_html = ''
        has_data = False

        for score_name, pred_name, label in PREDICTION_SCORE_NAMES:
            score = self._val_n(row_idx, score_name)
            pred = self._val_n(row_idx, pred_name) if pred_name else ""

            if not score and not pred:
                continue

            has_data = True
            score_display = escape(score) if score else "-"

            if pred:
                badge_cls = _pred_badge_class(pred)
                pred_display = f'<span class="pred-badge {badge_cls}">{escape(pred)}</span>'
            else:
                pred_display = "-"

            rows_html += f'''
                <tr>
                    <td style="font-weight:600;">{escape(label)}</td>
                    <td style="font-family:monospace;">{score_display}</td>
                    <td>{pred_display}</td>
                </tr>'''

        if not has_data:
            return '<div style="padding:10px;color:var(--text-muted);font-style:italic;">No prediction data available</div>'

        return f'''
            <table class="score-table">
                <thead>
                    <tr><th>Tool</th><th>Score</th><th>Prediction</th></tr>
                </thead>
                <tbody>{rows_html}
                </tbody>
            </table>'''

    def _pop_freq_html(self, row_idx: int) -> str:
        """Generate population frequency section, skipping empty values"""
        items = []
        for col_idx in self._cols(COLUMN_GROUP_NAMES["population_freq"]):
            value = self._val(row_idx, col_idx)
            if not value:
                continue
            col_name = self._header(col_idx)
            label = escape(POPULATION_FREQ_LABELS.get(col_name, col_name))

            # Try to add frequency bar
            bar_html = ""
            try:
                freq = float(value)
                pct = min(freq * 100, 100)
                bar_class = "rare" if freq < 0.01 else ("low" if freq < 0.05 else "common")
                bar_html = f'''
                            <div class="freq-bar-container">
                                <div class="freq-bar {bar_class}" style="width: {pct}%"></div>
                            </div>'''
            except (ValueError, TypeError):
                pass

            items.append(f'''
                    <div class="data-item">
                        <div class="data-label">{label}</div>
                        <div class="data-value monospace">{escape(value)}</div>{bar_html}
                    </div>''')

        if not items:
            return '<div style="padding:10px;color:var(--text-muted);font-style:italic;">No population frequency data available</div>'
        return '<div class="data-grid">' + ''.join(items) + '\n</div>'

    # ---- Page generators ----

    def generate_sample_page(self, sample_name: str) -> str:
        row_indices = self.samples[sample_name]
        safe_sample = re.sub(r'[^\w\-_]', '_', sample_name)

        # Build 0-based col_map for the clinical summary / card helpers.
        # self.col_idx stores 1-based indices; helpers expect 0-based direct row access.
        col_map_0 = {name: (idx - 1) for name, idx in self.col_idx.items()}

        # Collect raw row data lists for the clinical helpers.
        variants = [self.rows[idx] for idx in row_indices]

        # Build Clinical Summary tab content.
        # sample_html_dir is the directory that will contain the sample HTML file;
        # used to compute correct relative paths to IGV snapshots.
        sample_html_dir = self.output_dir / "samples"
        clinical_table = self._clinical_summary_table_html(variants, col_map_0, sample_html_dir)

        html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Variant Report - {escape(sample_name)}</title>
    <style>{get_css_styles()}</style>
    <script>{get_javascript()}</script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Variant Analysis Report</h1>
            <div class="subtitle">Sample: {escape(sample_name)}</div>
        </div>

        <div class="breadcrumb">
            <a href="../Summary.html">&#8592; All Samples</a> &gt; {escape(sample_name)}
        </div>

        <div class="dashboard-stats">
            <div class="stat-card">
                <div class="stat-value">{len(row_indices)}</div>
                <div class="stat-label">Total Variants</div>
            </div>'''

        genes = set()
        for idx in row_indices:
            g = self._val_n(idx, KEY_COL_NAMES["gene"])
            if g:
                genes.add(g)

        html += f'''
            <div class="stat-card">
                <div class="stat-value" style="color: var(--secondary-color);">{len(genes)}</div>
                <div class="stat-label">Genes Affected</div>
            </div>
        </div>

        <div class="tab-bar">
          <button class="tab-btn active" onclick="switchTab(this, \'tab-clinical\')">Clinical Summary</button>
          <button class="tab-btn" onclick="switchTab(this, \'tab-full\')">Full Annotation</button>
        </div>

        <div id="tab-clinical" class="tab-content">
          {clinical_table}
        </div>

        <div id="tab-full" class="tab-content" style="display:none">

        <table class="variant-table">
            <thead>
                <tr>
                    <th>#</th>
                    <th>Gene</th>
                    <th>Variant</th>
                    <th>HGVSc</th>
                    <th>HGVSp</th>
                    <th>VAF</th>
                    <th>DP</th>
                    <th>Action</th>
                </tr>
            </thead>
            <tbody>'''

        for i, row_idx in enumerate(row_indices):
            gene = self._val_n(row_idx, KEY_COL_NAMES["gene"]) or "-"
            chrom = self._val_n(row_idx, KEY_COL_NAMES["chr"]) or "-"
            pos = self._val_n(row_idx, KEY_COL_NAMES["pos"]) or "-"
            ref = self._val_n(row_idx, KEY_COL_NAMES["ref"]) or "-"
            alt = self._val_n(row_idx, KEY_COL_NAMES["alt"]) or "-"
            hgvsc = self._val_n(row_idx, KEY_COL_NAMES["hgvsc"]) or "-"
            hgvsp = self._val_n(row_idx, KEY_COL_NAMES["hgvsp"]) or "-"
            vaf_raw = self._val_n(row_idx, KEY_COL_NAMES["vaf"]) or "-"
            dp = self._val_n(row_idx, KEY_COL_NAMES["dp"]) or "-"

            vaf_display = vaf_raw
            vaf_bar_html = ""
            try:
                vaf_num = float(vaf_raw)
                vaf_pct = vaf_num * 100 if vaf_num <= 1 else vaf_num
                vaf_display = f"{vaf_pct:.1f}%"
                vaf_bar_html = (
                    f'<div class="vaf-bar-mini">'
                    f'<div class="vaf-bar-mini-fill" style="width:{min(vaf_pct,100):.1f}%"></div>'
                    f'</div>'
                )
            except (ValueError, TypeError):
                pass

            hgvsc_short = hgvsc[:30] + "..." if len(hgvsc) > 30 else hgvsc
            hgvsp_short = hgvsp[:20] + "..." if len(hgvsp) > 20 else hgvsp

            html += f'''
                <tr>
                    <td>{i+1}</td>
                    <td class="gene">{escape(gene)}</td>
                    <td style="font-family:monospace;">{escape(chrom)}:{escape(pos)} {escape(ref)}&gt;{escape(alt)}</td>
                    <td style="font-family:monospace;" title="{escape(hgvsc)}">{escape(hgvsc_short)}</td>
                    <td style="font-family:monospace;" title="{escape(hgvsp)}">{escape(hgvsp_short)}</td>
                    <td class="vaf-cell">{escape(vaf_display)}{vaf_bar_html}</td>
                    <td>{escape(dp)}</td>
                    <td><a href="variants/{safe_sample}_var{row_idx}.html">View Details</a></td>
                </tr>'''

        html += '''
            </tbody>
        </table>
        </div><!-- end tab-full -->
    </div>
</body>
</html>'''
        return html

    def generate_variant_page(self, sample_name: str, row_idx: int) -> str:
        safe_sample = re.sub(r'[^\w\-_]', '_', sample_name)

        gene = self._val_n(row_idx, KEY_COL_NAMES["gene"]) or "Unknown"
        chrom = self._val_n(row_idx, KEY_COL_NAMES["chr"]) or "-"
        pos = self._val_n(row_idx, KEY_COL_NAMES["pos"]) or "-"
        ref = self._val_n(row_idx, KEY_COL_NAMES["ref"]) or "-"
        alt = self._val_n(row_idx, KEY_COL_NAMES["alt"]) or "-"
        hgvsc = self._val_n(row_idx, KEY_COL_NAMES["hgvsc"]) or "-"
        hgvsp = self._val_n(row_idx, KEY_COL_NAMES["hgvsp"]) or "-"

        igv_path = self.find_igv_screenshot(sample_name, chrom, pos)
        variant_html_dir = (self.output_dir / "samples" / "variants").resolve()

        html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{escape(gene)} {escape(hgvsp)} - {escape(sample_name)}</title>
    <style>{get_css_styles()}</style>
    <script>{get_javascript()}</script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Variant Analysis Report</h1>
            <div class="subtitle">Sample: {escape(sample_name)}</div>
        </div>

        <div class="breadcrumb">
            <a href="../../Summary.html">All Samples</a> &gt;
            <a href="../{safe_sample}.html">{escape(sample_name)}</a> &gt;
            {escape(gene)} {escape(chrom)}:{escape(pos)}
        </div>

        <div class="gene-hero">
            <div>
                <h1 class="gene-name">{escape(gene)}</h1>
                <div class="variant-notation">{escape(hgvsp) if hgvsp != "-" else escape(hgvsc)}</div>
            </div>
            {f'<div>{_cancervar_tier_badge(self._val_n(row_idx, "CancerVar and Evidence"))}</div>'}
        </div>'''

        # 1. Basic Variant Information
        html += self._panel_html(
            "Basic Variant Information", "panel-basic",
            self._data_grid_html(row_idx, self._cols(COLUMN_GROUP_NAMES["basic_info"])),
            collapsed=False
        )

        # 2. IGV Screenshot (after Basic Info)
        if igv_path:
            rel_igv = os.path.relpath(igv_path, str(variant_html_dir))
            igv_content = f'''
                <img src="{rel_igv}" alt="IGV Screenshot for {escape(gene)} at {escape(chrom)}:{escape(pos)}"
                     class="igv-image" onerror="this.style.display='none'; this.nextElementSibling.style.display='block';">
                <div class="igv-no-image" style="display: none;">
                    Screenshot not available
                </div>'''
        else:
            igv_content = '<div class="igv-no-image">No IGV screenshot available for this variant</div>'

        html += self._panel_html("IGV Screenshot", "panel-igv", igv_content, collapsed=False)

        # 3. Sample Comparison
        html += self._panel_html(
            "Sample Comparison Data", "panel-comparison",
            self._data_grid_html(row_idx, self._cols(COLUMN_GROUP_NAMES["sample_comparison"])),
            collapsed=False
        )

        # 4. Additional Variant Information
        html += self._panel_html(
            "Additional Variant Information", "panel-additional",
            self._data_grid_html(row_idx, self._cols(COLUMN_GROUP_NAMES["additional_info"])),
            collapsed=False
        )

        # 5. Population Frequency Databases
        html += self._panel_html(
            "Population Frequency Databases", "panel-population",
            self._pop_freq_html(row_idx),
            collapsed=True
        )

        # 6. ACMG Classification Criteria
        html += self._panel_html(
            "ACMG Classification Criteria", "panel-acmg",
            self._acmg_section_html(row_idx),
            collapsed=True
        )

        # 7. Computational Predictions
        html += self._panel_html(
            "Computational Predictions", "panel-computational",
            self._computational_section_html(row_idx),
            collapsed=True
        )

        html += '''
    </div>
</body>
</html>'''
        return html

    def generate(self):
        """End-to-end entry point: load Excel data then write all HTML reports."""
        self.load_excel()
        self.generate_reports()

    def generate_reports(self):
        self.output_dir.mkdir(parents=True, exist_ok=True)
        samples_dir = self.output_dir / "samples"
        samples_dir.mkdir(exist_ok=True)
        variants_dir = samples_dir / "variants"
        variants_dir.mkdir(exist_ok=True)

        # Copy snapshots inside html_reports/ so they're accessible under the web server root
        if self.snapshots_dir and self.snapshots_dir.exists():
            import shutil
            dest_snapshots = self.output_dir / "SnapShots"
            if dest_snapshots.resolve() != self.snapshots_dir.resolve():
                if dest_snapshots.exists():
                    shutil.rmtree(str(dest_snapshots))
                shutil.copytree(str(self.snapshots_dir), str(dest_snapshots))
            self.snapshots_dir = dest_snapshots

        for sample_name in sorted(self.samples.keys()):
            safe_sample = re.sub(r'[^\w\-_]', '_', sample_name)
            print(f"Generating pages for sample: {sample_name}")
            (samples_dir / f"{safe_sample}.html").write_text(self.generate_sample_page(sample_name))

            for row_idx in self.samples[sample_name]:
                (variants_dir / f"{safe_sample}_var{row_idx}.html").write_text(
                    self.generate_variant_page(sample_name, row_idx))

        print(f"\nHTML reports generated in: {self.output_dir}")
        print(f"  - Sample pages: {samples_dir}")
        print(f"  - Variant pages: {variants_dir}")


def generate_summary_page(html_dir: str):
    """Generate Summary.html landing page from existing sample HTML files."""
    html_path = Path(html_dir)
    samples_dir = html_path / "samples"

    # Collect sample names from existing HTML files
    sample_files = sorted(samples_dir.glob("*.html")) if samples_dir.exists() else []

    date_str = ""
    try:
        from datetime import date
        date_str = date.today().strftime("%Y-%m-%d")
    except Exception:
        pass

    variants_dir = samples_dir / "variants"
    cards_html = ""
    total_variants = 0
    for sf in sample_files:
        sample_name = sf.stem
        n_variants = len(list(variants_dir.glob(f"{sample_name}_var*.html"))) if variants_dir.exists() else 0
        total_variants += n_variants
        card_class = "sample-card no-variants" if n_variants == 0 else "sample-card"
        no_variants_badge = '<div class="no-variants-badge">No variants</div>' if n_variants == 0 else ""
        count_class = "sample-stat-value zero" if n_variants == 0 else "sample-stat-value"

        # Determine clinical signal tier by scanning sample HTML for badge classes
        data_tier = ""
        if n_variants > 0:
            try:
                sample_html_content = sf.read_text(errors="replace")
                if "badge-pathogenic" in sample_html_content:
                    data_tier = ' data-tier="1"'
                elif "badge-likely-pathogenic" in sample_html_content:
                    data_tier = ' data-tier="2"'
                else:
                    data_tier = ' data-tier="3"'
            except Exception:
                data_tier = ' data-tier="3"'

        cards_html += f'''
            <a href="samples/{escape(sf.name)}" class="{card_class}"{data_tier}>
                <div class="sample-card-header">
                    <h3>{escape(sample_name)}</h3>
                    {no_variants_badge}
                </div>
                <div class="sample-card-body">
                    <div class="sample-stat">
                        <span class="sample-stat-label">Variants</span>
                        <span class="{count_class}">{n_variants}</span>
                    </div>
                </div>
            </a>'''

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>iSeq Variant Analysis Summary</title>
    <style>{get_css_styles()}</style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>iSeq Variant Analysis Summary</h1>
            <div class="subtitle">Generated: {date_str} &mdash; {len(sample_files)} sample(s)</div>
        </div>
        <div class="dashboard-stats">
            <div class="stat-card">
                <div class="stat-value">{len(sample_files)}</div>
                <div class="stat-label">Samples</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{total_variants}</div>
                <div class="stat-label">Total Variants</div>
            </div>
        </div>
        <div class="sample-grid">
            {cards_html}
        </div>
    </div>
</body>
</html>'''

    summary_file = html_path / "Summary.html"
    summary_file.write_text(html)
    print(f"Generated: {summary_file}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 excel_to_html_report.py <input_excel_file> [output_directory] [snapshots_directory]")
        sys.exit(1)

    if sys.argv[1] == "--summary":
        if len(sys.argv) < 3:
            print("Usage: python3 excel_to_html_report.py --summary <html_reports_directory>")
            sys.exit(1)
        generate_summary_page(sys.argv[2])
        return

    excel_path = sys.argv[1]
    if not os.path.isfile(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    output_dir = sys.argv[2] if len(sys.argv) > 2 else "html_report"
    snapshots_dir = sys.argv[3] if len(sys.argv) > 3 else None

    generator = iSeqReportGenerator(excel_path, output_dir, snapshots_dir)
    generator.load_excel()
    generator.generate_reports()


if __name__ == "__main__":
    main()
